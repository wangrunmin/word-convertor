#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
筛查结果导出器 export_screening_report.py
读取 screening.sqlite，解析原始接口返回，统一导出为 Excel（失败降级 CSV）。

用法：
  python export_screening_report.py                          # 交互向导
  python export_screening_report.py -d screen_out/screening.sqlite -i json/ -o report.xlsx
  python export_screening_report.py --rescan-meta -i json/  # 强制重扫法院/案号
"""

import argparse
import csv
import json
import os
import re
import sqlite3
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, Generator, List, Optional, Tuple

# ── 可选依赖 ──────────────────────────────────────────────────────────────────

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from tqdm import tqdm as _tqdm
    def tqdm(it, desc="", **kw):
        return _tqdm(it, desc=desc, **kw)
except ImportError:
    def tqdm(it, desc="", **kw):
        items = list(it)
        print(f"  {desc}（共 {len(items)} 项）…", flush=True)
        return items

try:
    import yaml
    HAS_YAML = True
except ImportError:
    HAS_YAML = False

# ── 常量 ──────────────────────────────────────────────────────────────────────

CONFIG_FILE       = "screening_config.yaml"
CONFIG_FILE_LOCAL = "screening_config.local.yaml"
DB_DEFAULT        = Path("screen_out") / "screening.sqlite"
OUT_DEFAULT_STEM  = "筛查报告"

ISSUE_COLUMNS = [
    "文件", "法院", "案号", "接口", "模块",
    "一级分类", "二级分类", "三级分类", "维度枚举码", "描述",
    "法律名称", "条文", "建议表述",
    "关键词", "修改建议", "原文", "纠错依据", "定位",
]

SUMMARY_COLUMNS = [
    "文件", "法院", "案号", "接口", "状态", "HTTP状态码",
    "重试次数", "耗时(ms)", "采集时间", "问题数", "错误信息",
]

FAILED_COLUMNS = [
    "文件", "法院", "案号", "接口", "状态", "HTTP状态码",
    "重试次数", "采集时间", "错误信息",
]

# xlsx 颜色
HDR_FILL = "2F5496"
HDR_FONT = "FFFFFF"
ALT_FILL = "DCE6F1"
OK_FILL  = "E2EFDA"
ERR_FILL = "FCE4D6"

WRAP_COLS = {"原文", "纠错依据", "修改建议"}

# ── 法院 / 案号提取（宽松正则，复用参考脚本经验证的规则）──────────────────────

_RE_CASE_NUM = re.compile(r'[（【]\d{4}[）】][^\s\r\n（）【】]{1,30}号?')
_RE_COURT    = re.compile(r'[一-鿿　（）·•]{2,40}(?:法院|检察院|仲裁委员会|仲裁院|仲裁委|法庭)')


def extract_court_and_case(paragraphs: list) -> Tuple[str, str]:
    court = ""
    case_num = ""
    first_nonempty = ""
    for para in paragraphs[:20]:
        raw  = para.get("content") or ""
        text = raw.replace("\r", "").replace("\n", "").strip()
        if not text:
            continue
        if not first_nonempty:
            first_nonempty = text
        if not case_num:
            m = _RE_CASE_NUM.search(raw)
            if m:
                case_num = m.group().strip()
        if not court:
            m = _RE_COURT.search(text)
            if m:
                court = m.group()
        if court and case_num:
            break
    if not court:
        court = first_nonempty
    return court, case_num


# ── Parser registry ───────────────────────────────────────────────────────────

PARSERS: Dict[str, callable] = {}


def parser(name: str):
    def deco(fn):
        PARSERS[name] = fn
        return fn
    return deco


@parser("keypoint")
def parse_keypoint(raw: dict, row: dict) -> Generator:
    """重点纠错 / 风险防控：解析 corrections[] 列表"""
    for item in (raw or {}).get("corrections") or []:
        ct  = item.get("correctionTypes") or {}
        idx = (item.get("indexs") or [{}])[0]
        yield {
            "文件":     row["doc_path"],
            "法院":     row["court"],
            "案号":     row["case_num"],
            "接口":     row["interface"],
            "模块":     ct.get("module", "纠错"),
            "一级分类": ct.get("errorcorrectionType", ""),
            "二级分类": ct.get("category", ""),
            "三级分类": ct.get("errorClassification", ""),
            "维度枚举码": ct.get("type", ""),
            "描述":     ct.get("description", ""),
            "法律名称": "",
            "条文":     "",
            "建议表述": "",
            "关键词":   item.get("keyword", ""),
            "修改建议": item.get("suggestion", ""),
            "原文":     idx.get("originalText") or idx.get("pcOriginalText", ""),
            "纠错依据": item.get("tip", ""),
            "定位":     idx.get("index", ""),
        }


@parser("law_quote")
def parse_law_quote(raw: dict, row: dict) -> Generator:
    """法律纠错 / 法条引用：解析 lawQuotes[] 列表"""
    for item in (raw or {}).get("lawQuotes") or []:
        idx       = (item.get("indexs") or [{}])[0]
        o_law     = item.get("oLawName", "")
        tiao      = item.get("tiao", "")
        law       = item.get("lawName", "")
        sugg_tiao = item.get("suggestedTiao") or ""
        yield {
            "文件":     row["doc_path"],
            "法院":     row["court"],
            "案号":     row["case_num"],
            "接口":     row["interface"],
            "模块":     "法律纠错",
            "一级分类": item.get("categoryStr", ""),
            "二级分类": item.get("categoryStr", ""),
            "三级分类": item.get("subCategory", ""),
            "维度枚举码": item.get("category", ""),
            "描述":     item.get("tip", ""),
            "法律名称": o_law,
            "条文":     tiao,
            "建议表述": sugg_tiao,
            "关键词":   (o_law + tiao).strip(),
            "修改建议": (law + sugg_tiao).strip() or item.get("proposal", ""),
            "原文":     idx.get("pcOriginalText", ""),
            "纠错依据": item.get("tip", ""),
            "定位":     idx.get("index", ""),
        }


# ── SQLite helpers ─────────────────────────────────────────────────────────────

def open_db(db_path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS document_meta (
            doc_path   TEXT PRIMARY KEY,
            court      TEXT,
            case_num   TEXT,
            scanned_at TEXT NOT NULL
        )
    """)
    conn.commit()
    return conn


def fill_document_meta(conn: sqlite3.Connection, doc_paths: List[str],
                       input_dir: Path, force: bool = False):
    if force:
        todo = list(doc_paths)
    else:
        existing = {r[0] for r in conn.execute("SELECT doc_path FROM document_meta")}
        todo = [p for p in doc_paths if p not in existing]

    if not todo:
        print(f"  ✓ 元数据缓存已是最新（{len(doc_paths)} 篇，无需重扫）")
        return

    print(f"  需要扫描 {len(todo)} 篇文档（已缓存 {len(doc_paths) - len(todo)} 篇）")
    now   = datetime.now().isoformat(timespec="seconds")
    sql   = ("INSERT OR REPLACE INTO document_meta(doc_path, court, case_num, scanned_at) "
             "VALUES (?, ?, ?, ?)")
    batch: list = []
    for rel in tqdm(todo, desc="提取法院/案号"):
        try:
            data = json.loads((input_dir / Path(rel)).read_text(encoding="utf-8"))
            court, case_num = extract_court_and_case(data.get("paragraphs") or [])
        except Exception:
            court, case_num = "", ""
        batch.append((rel, court, case_num, now))
        if len(batch) >= 500:
            conn.executemany(sql, batch)
            batch.clear()
    if batch:
        conn.executemany(sql, batch)
    conn.commit()
    print(f"  ✓ 元数据缓存已更新（{len(todo)} 篇）")


def query_rows(conn: sqlite3.Connection,
               ifaces: Optional[List[str]],
               status_filter: Optional[str]) -> list:
    params: list = []
    where:  list = []
    if ifaces:
        ph = ",".join("?" * len(ifaces))
        where.append(f"r.interface IN ({ph})")
        params.extend(ifaces)
    if status_filter:
        where.append("r.status = ?")
        params.append(status_filter)
    clause = ("WHERE " + " AND ".join(where)) if where else ""
    sql = f"""
        SELECT r.doc_path, r.interface, r.status,
               r.raw_response, r.error, r.http_status,
               r.attempts, r.elapsed_ms, r.fetched_at,
               COALESCE(m.court,    '') AS court,
               COALESCE(m.case_num, '') AS case_num
          FROM screening_result r
          LEFT JOIN document_meta m ON m.doc_path = r.doc_path
         {clause}
         ORDER BY r.doc_path, r.interface
    """
    return conn.execute(sql, params).fetchall()


# ── xlsx 样式 helpers ─────────────────────────────────────────────────────────

def _fill(hex_color: str) -> "PatternFill":
    return PatternFill("solid", fgColor=hex_color)


def _style_header(ws, headers: list):
    hf = _fill(HDR_FILL)
    hfont = Font(color=HDR_FONT, bold=True)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.fill  = hf
        cell.font  = hfont
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _autowidth(ws, headers: list):
    widths = [len(h) for h in headers]
    for row in ws.iter_rows(min_row=2):
        for ci, cell in enumerate(row):
            if ci < len(widths) and cell.value:
                widths[ci] = max(widths[ci], min(len(str(cell.value)), 80))
    for ci, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(ci + 1)].width = max(w + 2, 8)


def _row_styles(ws, headers: list, status_col: Optional[str] = None,
                fill_all: Optional[str] = None):
    alt  = _fill(ALT_FILL)
    ok   = _fill(OK_FILL)
    err  = _fill(ERR_FILL)
    forced = _fill(fill_all) if fill_all else None
    wrap_align   = Alignment(wrap_text=True,  vertical="top")
    normal_align = Alignment(wrap_text=False, vertical="top")
    sidx = (headers.index(status_col) + 1) if status_col and status_col in headers else None
    for ri, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if forced:
            fill = forced
        else:
            sv = str(row[sidx - 1].value or "") if sidx else ""
            if sv == "success":
                fill = ok
            elif sv == "failed":
                fill = err
            elif ri % 2 == 0:
                fill = alt
            else:
                fill = None
        for ci, cell in enumerate(row, 1):
            col_name = headers[ci - 1] if ci <= len(headers) else ""
            cell.alignment = wrap_align if col_name in WRAP_COLS else normal_align
            if fill:
                cell.fill = fill


# ── 四个 sheet ────────────────────────────────────────────────────────────────

def write_overview(ws, summary_rows: list, issues: list, elapsed: float):
    ws.title = "总览"
    bold = Font(bold=True)
    hf   = _fill(HDR_FILL)
    hfont = Font(color=HDR_FONT, bold=True)

    def kv(row, key, val):
        c = ws.cell(row, 1, key)
        c.font = bold
        ws.cell(row, 2, val)

    kv(1, "导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    kv(2, "文档总数", len({r["doc_path"] for r in summary_rows}))
    kv(3, "问题明细总数", len(issues))
    kv(4, "总耗时(s)", round(elapsed, 1))

    # 接口统计表头
    for ci, h in enumerate(["接口", "成功", "失败", "问题数"], 1):
        cell = ws.cell(6, ci, h)
        cell.fill = hf
        cell.font = hfont

    iface_stats: Dict[str, Dict] = {}
    for r in summary_rows:
        s = iface_stats.setdefault(r["interface"], {"success": 0, "failed": 0, "issues": 0})
        if r["status"] == "success":
            s["success"] += 1
        else:
            s["failed"] += 1
    for issue in issues:
        iface_stats.setdefault(issue["接口"], {"success": 0, "failed": 0, "issues": 0})["issues"] += 1

    for ri, (iname, s) in enumerate(sorted(iface_stats.items()), 7):
        ws.cell(ri, 1, iname)
        ws.cell(ri, 2, s["success"])
        ws.cell(ri, 3, s["failed"])
        ws.cell(ri, 4, s["issues"])

    for col, w in zip("ABCD", [20, 12, 12, 12]):
        ws.column_dimensions[col].width = w


def write_summary(ws, summary_rows: list):
    ws.title = "筛查结果"
    _style_header(ws, SUMMARY_COLUMNS)
    for r in summary_rows:
        ws.append([
            r["doc_path"], r["court"], r["case_num"],
            r["interface"], r["status"], r.get("http_status"),
            r.get("attempts"), r.get("elapsed_ms"), r.get("fetched_at"),
            r.get("issue_count", 0), r.get("error") or "",
        ])
    _row_styles(ws, SUMMARY_COLUMNS, status_col="状态")
    _autowidth(ws, SUMMARY_COLUMNS)


def write_issues(ws, issues: list):
    ws.title = "问题明细"
    _style_header(ws, ISSUE_COLUMNS)
    for issue in issues:
        ws.append([issue.get(c, "") for c in ISSUE_COLUMNS])
    _row_styles(ws, ISSUE_COLUMNS)
    _autowidth(ws, ISSUE_COLUMNS)


def write_failed(ws, failed_rows: list):
    ws.title = "失败列表"
    _style_header(ws, FAILED_COLUMNS)
    for r in failed_rows:
        ws.append([
            r["doc_path"], r.get("court", ""), r.get("case_num", ""),
            r["interface"], r["status"], r.get("http_status"),
            r.get("attempts"), r.get("fetched_at"), r.get("error") or "",
        ])
    _row_styles(ws, FAILED_COLUMNS, fill_all=ERR_FILL)
    _autowidth(ws, FAILED_COLUMNS)


# ── 写入 xlsx ─────────────────────────────────────────────────────────────────

def do_write_xlsx(out_path: Path, summary_rows: list, issues: list,
                  failed_rows: list, elapsed: float) -> Path:
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    target = out_path.parent / f"{out_path.stem}_{ts}.xlsx"
    target.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)
    write_overview(wb.create_sheet("总览"),       summary_rows, issues, elapsed)
    write_summary( wb.create_sheet("筛查结果"),   summary_rows)
    write_issues(  wb.create_sheet("问题明细"),   issues)
    if failed_rows:
        write_failed(wb.create_sheet("失败列表"), failed_rows)

    try:
        wb.save(str(target))
        return target
    except PermissionError:
        retry = out_path.parent / f"{out_path.stem}_{ts}_retry.xlsx"
        wb.save(str(retry))
        return retry


# ── 降级 CSV ──────────────────────────────────────────────────────────────────

def do_write_csv(base_dir: Path, summary_rows: list, issues: list,
                 failed_rows: list) -> Path:
    ts  = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    out = base_dir / f"report_csv_{ts}"
    out.mkdir(parents=True, exist_ok=True)

    def write_csv(name: str, headers: list, rows_iter):
        with open(out / f"{name}.csv", "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
            w.writerow(headers)
            for r in rows_iter:
                w.writerow(r)

    write_csv("筛查结果", SUMMARY_COLUMNS, (
        [r["doc_path"], r.get("court", ""), r.get("case_num", ""),
         r["interface"], r["status"], r.get("http_status"),
         r.get("attempts"), r.get("elapsed_ms"), r.get("fetched_at"),
         r.get("issue_count", 0), r.get("error") or ""]
        for r in summary_rows
    ))
    write_csv("问题明细", ISSUE_COLUMNS, (
        [issue.get(c, "") for c in ISSUE_COLUMNS] for issue in issues
    ))
    if failed_rows:
        write_csv("失败列表", FAILED_COLUMNS, (
            [r["doc_path"], r.get("court", ""), r.get("case_num", ""),
             r["interface"], r["status"], r.get("http_status"),
             r.get("attempts"), r.get("fetched_at"), r.get("error") or ""]
            for r in failed_rows
        ))
    return out


# ── 配置加载 ──────────────────────────────────────────────────────────────────

def resolve_config_path() -> Path:
    local = Path(CONFIG_FILE_LOCAL)
    return local if local.exists() else Path(CONFIG_FILE)


def load_iface_parser_map() -> Dict[str, Optional[str]]:
    cfg_path = resolve_config_path()
    if not cfg_path.exists() or not HAS_YAML:
        return {}
    try:
        cfg = yaml.safe_load(cfg_path.read_text(encoding="utf-8")) or {}
        return {k: v.get("parser") for k, v in cfg.get("interfaces", {}).items()}
    except Exception as e:
        print(f"  ⚠ 配置文件读取失败: {e}", file=sys.stderr)
        return {}


# ── 主导出逻辑 ────────────────────────────────────────────────────────────────

def run_export(
    db_path:       Path,
    input_dir:     Optional[Path],
    out_path:      Path,
    ifaces:        Optional[List[str]],
    status_filter: Optional[str],
    fmt:           str,
    rescan_meta:   bool = False,
) -> Path:
    t0 = datetime.now()

    iface_parser_map = load_iface_parser_map()
    if not iface_parser_map:
        print("  ⚠ 未加载到接口配置，所有接口跳过问题明细解析")
    else:
        for k, v in iface_parser_map.items():
            if not v:
                print(f"  ⚠ 接口 {k} 未配置 parser，跳过问题明细解析")

    conn = open_db(db_path)

    # 元数据缓存填充
    if input_dir and input_dir.exists():
        doc_paths = [r[0] for r in conn.execute(
            "SELECT DISTINCT doc_path FROM screening_result"
        )]
        fill_document_meta(conn, doc_paths, input_dir, force=rescan_meta)
    elif rescan_meta:
        print("  ⚠ --rescan-meta 需要 -i 指定 JSON 目录，跳过")
    elif not input_dir:
        meta_count = conn.execute("SELECT COUNT(*) FROM document_meta").fetchone()[0]
        if meta_count == 0:
            print("  ⚠ 未指定 -i 且无元数据缓存，法院/案号列将为空")

    # 查询
    rows = query_rows(conn, ifaces, status_filter)
    print(f"  共 {len(rows)} 行待处理")

    issues:  List[Dict] = []
    summary: List[Dict] = []
    failed:  List[Dict] = []
    parse_err = 0

    for r in rows:
        r_dict = dict(r)
        if r_dict["status"] == "failed":
            failed.append(r_dict)

        ptype   = iface_parser_map.get(r_dict["interface"])
        n_issue = 0
        if ptype and ptype in PARSERS:
            try:
                raw_resp = r_dict.get("raw_response")
                raw_json = json.loads(raw_resp) if raw_resp else {}
                for issue in PARSERS[ptype](raw_json, r_dict):
                    issues.append(issue)
                    n_issue += 1
            except Exception as e:
                parse_err += 1
                print(
                    f"  ⚠ 解析失败 {r_dict['doc_path']} [{r_dict['interface']}]: {e}",
                    file=sys.stderr,
                )

        r_dict["issue_count"] = n_issue
        summary.append(r_dict)

    if parse_err:
        print(f"  ⚠ 共 {parse_err} 行解析失败")
    print(f"  ✓ 解析完成：{len(issues)} 条问题明细，{len(failed)} 条失败行")

    elapsed = (datetime.now() - t0).total_seconds()

    # 写出
    if fmt == "xlsx" and not HAS_OPENPYXL:
        print("  ⚠ openpyxl 未安装（pip install openpyxl），自动降级为 CSV")
        fmt = "csv"

    if fmt == "xlsx":
        try:
            result_path = do_write_xlsx(out_path, summary, issues, failed, elapsed)
            print(f"\n  ✓ 已写入: {result_path}")
        except Exception as e:
            print(f"  ✗ xlsx 写入失败: {e}，降级 CSV", file=sys.stderr)
            result_path = do_write_csv(out_path.parent, summary, issues, failed)
            print(f"\n  ✓ 已写入 CSV 目录: {result_path}")
    else:
        result_path = do_write_csv(out_path.parent, summary, issues, failed)
        print(f"\n  ✓ 已写入 CSV 目录: {result_path}")

    conn.close()
    return result_path


def open_result(path: Path):
    try:
        if sys.platform == "win32":
            os.startfile(str(path))
        elif sys.platform == "darwin":
            os.system(f'open "{path}"')
        else:
            os.system(f'xdg-open "{path}"')
    except Exception:
        pass


# ── 交互向导 ──────────────────────────────────────────────────────────────────

def strip_path(s: str) -> str:
    return s.strip().strip('"').strip("'").strip()


def run_wizard():
    print()
    print("══════════════════════════════════════════")
    print("   筛查结果导出器 v1")
    print("══════════════════════════════════════════")
    print()

    # 1/5 SQLite
    print(f" 1/5  SQLite 路径（阶段三输出，默认: {DB_DEFAULT}）")
    print("      提示：可直接拖拽文件到窗口")
    raw = input("      路径: ").strip()
    db_path = Path(strip_path(raw)) if raw else DB_DEFAULT
    if not db_path.exists():
        print(f"  ✗ 找不到文件: {db_path}")
        sys.exit(1)
    conn_tmp = sqlite3.connect(str(db_path))
    total = conn_tmp.execute("SELECT COUNT(*) FROM screening_result").fetchone()[0]
    conn_tmp.close()
    print(f"      ✓ 找到，共 {total} 行\n")

    # 2/5 JSON 目录（可选）
    print(" 2/5  阶段二 JSON 目录（提取法院/案号用，可留空）")
    print("      留空则跳过（已有缓存的会直接复用）")
    raw = input("      路径: ").strip()
    input_dir = Path(strip_path(raw)) if raw else None

    # 3/5 导出范围
    print("\n 3/5  导出范围")
    print("      [1] 全部（默认）")
    print("      [2] 仅成功")
    print("      [3] 仅失败")
    choice = input("      选择: ").strip() or "1"
    status_filter = None if choice == "1" else ("success" if choice == "2" else "failed")

    # 4/5 输出路径
    default_out = db_path.parent / f"{OUT_DEFAULT_STEM}.xlsx"
    print(f"\n 4/5  输出路径（默认: {default_out}）")
    raw = input("      路径: ").strip()
    out_path = Path(strip_path(raw)) if raw else default_out

    # 5/5 格式
    print("\n 5/5  输出格式")
    print("      [1] xlsx（推荐，默认）")
    print("      [2] csv（多文件，兼容所有 Excel）")
    choice = input("      选择: ").strip() or "1"
    fmt = "csv" if choice == "2" else "xlsx"

    print()
    return db_path, input_dir, out_path, None, status_filter, fmt


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) == 1:
        db_path, input_dir, out_path, ifaces, status_filter, fmt = run_wizard()
        result = run_export(db_path, input_dir, out_path, ifaces, status_filter, fmt)
        ans = input("  是否打开报告？[Y/n]: ").strip().lower()
        if ans != "n":
            open_result(result)
        return

    p = argparse.ArgumentParser(
        description="筛查结果导出器：读 screening.sqlite，导出为 Excel/CSV",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="示例:\n"
               "  python export_screening_report.py                          # 向导\n"
               "  python export_screening_report.py -d screen_out/screening.sqlite -i json/ -o report.xlsx\n"
               "  python export_screening_report.py -i json/ --rescan-meta   # 刷新法院/案号缓存\n",
    )
    p.add_argument("-d", "--db",
                   default=str(DB_DEFAULT),
                   help=f"SQLite 路径（默认: {DB_DEFAULT}）")
    p.add_argument("-i", "--input",
                   help="阶段二 JSON 目录（用于填充/更新法院/案号缓存）")
    p.add_argument("-o", "--output",
                   help="输出文件路径（默认: <db目录>/筛查报告.xlsx）")
    p.add_argument("--interface",
                   help="逗号分隔的接口名子集，不传则全量（例如: A,B）")
    p.add_argument("--status",
                   choices=["success", "failed"],
                   help="按状态过滤，不传则全量")
    p.add_argument("--format",
                   choices=["xlsx", "csv"],
                   default="xlsx",
                   help="输出格式（默认: xlsx）")
    p.add_argument("--rescan-meta",
                   action="store_true",
                   help="强制重扫所有 JSON 的法院/案号（需要 -i）")
    args = p.parse_args()

    db_path   = Path(args.db)
    input_dir = Path(args.input) if args.input else None
    out_path  = Path(args.output) if args.output else db_path.parent / f"{OUT_DEFAULT_STEM}.xlsx"
    ifaces    = [x.strip() for x in args.interface.split(",")] if args.interface else None

    run_export(db_path, input_dir, out_path, ifaces, args.status,
               args.format, args.rescan_meta)


if __name__ == "__main__":
    main()
